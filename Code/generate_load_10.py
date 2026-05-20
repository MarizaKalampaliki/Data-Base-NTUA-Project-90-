from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
import os
import math
import datetime

EXCEL_FILE = "sample_data_10.xlsx"

EXCEL_FILE_ALTERNATIVE = "sample_data_10.xlsx"
OUTPUT_FOLDER = "sql"
DATABASE_NAME = "hospitaldb"

BATCH_SIZE = 150

TABLE_COLUMNS = {
    "staff": ["amka", "first_name", "last_name", "date_of_birth", "email", "phone", "hire_date", "staff_type"],
    "doctor": ["amka", "licence_number", "speciality", "rank", "supervisor_amka"],
    "department": ["department_id", "name", "description", "bed_count", "floor", "building", "head_doctor_amka"],
    "nurse": ["amka", "grade", "department_id"],
    "admin_staff": ["amka", "role", "department_id", "office"],
    "doctor_department": ["doctor_amka", "department_id"],
    "bed": ["bed_id", "bed_number", "type", "status", "department_id"],
    "insurance": ["provider_id", "insurance_number", "provider_name", "provider_type"],
    "patient": ["amka", "first_name", "last_name", "father_name", "date_of_birth", "gender", "weight", "height", "address", "phone", "email", "profession", "nationality", "provider_id"],
    "emergency_contact": ["contact_id", "patient_amka", "first_name", "last_name", "phone", "relationship"],
    "diagnosis": ["icd10_code", "description", "category_code", "category_name"],
    "ken": ["ken_code", "description", "base_cost", "expected_duration", "extra_daily_cost"],
    "triage": ["triage_id", "patient_amka", "nurse_amka", "symptoms", "priority_level", "arrival_time", "waiting_minutes", "outcome"],
    "hospitalization": ["hospitalization_id", "patient_amka", "department_id", "bed_id", "ken_code", "admission_date", "discharge_date", "admission_icd10_code", "discharge_icd10_code", "actual_duration", "total_cost", "triage_id"],
    "exam": ["exam_id", "hospitalization_id", "doctor_amka", "exam_code", "type", "exam_date", "result", "unit", "cost"],
    "operating_room": ["room_id", "room_number", "floor", "building", "type", "status"],
    "medical_procedure": ["procedure_id", "hospitalization_id", "room_id", "procedure_code", "name", "type", "procedure_date", "procedure_duration", "cost", "result", "head_doctor_amka"],
    "procedure_staff": ["procedure_id", "staff_amka", "role"],
    "drug": ["drug_id", "name", "form"],
    "active_substance": ["substance_id", "name"],
    "drug_active_substance": ["drug_id", "substance_id"],
    "patient_allergy": ["patient_amka", "substance_id"],
    "prescription": ["prescription_id", "hospitalization_id", "doctor_amka", "drug_id", "dosage", "frequency", "start_date", "end_date", "patient_amka"],
    "shift": ["shift_id", "department_id", "shift_date", "start_time", "end_time", "shift_type"],
    "shift_assignment": ["assignment_id", "shift_id", "staff_amka", "role", "status"],
    "hospitalization_evaluation": ["evaluation_id", "hospitalization_id", "evaluation_date", "nursing_care_score", "cleanliness_score", "food_score", "overall_experience_score", "comments"],
    "doctor_evaluation": ["doctor_evaluation_id", "hospitalization_id", "doctor_amka", "evaluation_date", "medical_care_score", "comments"],
    "entity_image": ["image_id", "entity_type", "entity_key", "image_path", "alt_text"],
}

DATE_COLUMNS = {
    "date_of_birth",
    "hire_date",
    "admission_date",
    "discharge_date",
    "start_date",
    "end_date",
    "shift_date",
    "evaluation_date",
}

DATETIME_COLUMNS = {"arrival_time", "exam_date", "procedure_date"}
TIME_COLUMNS = {"start_time", "end_time"}


def clean_name(value):
    text = str(value).strip().lower()

    for old, new in [
        (" ", "_"),
        ("-", "_"),
        ("/", "_"),
        (".", ""),
        (",", ""),
        (":", ""),
        (";", ""),
        ("(", ""),
        (")", ""),
    ]:
        text = text.replace(old, new)

    while "__" in text:
        text = text.replace("__", "_")

    return text


def is_blank(value):
    if value is None:
        return True

    if isinstance(value, float) and math.isnan(value):
        return True

    if isinstance(value, str) and value.strip() == "":
        return True

    return False


def to_datetime_value(value):
    if is_blank(value):
        return None

    if isinstance(value, datetime.datetime):
        return value

    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return from_excel(value)
        except Exception:
            return None

    text = str(value).strip()

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    ]

    for fmt in formats:
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            pass

    return None


def to_time_value(value):
    if is_blank(value):
        return None

    if isinstance(value, datetime.time):
        return value

    if isinstance(value, datetime.datetime):
        return value.time()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return from_excel(value).time()
        except Exception:
            return None

    text = str(value).strip()

    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            return datetime.datetime.strptime(text, fmt).time()
        except ValueError:
            pass

    return None


def sql_string(value):
    text = str(value)

    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = text.replace("[", "(")
    text = text.replace("]", ")")
    text = text.replace("\u202f", " ")
    text = text.replace("\u00a0", " ")
    text = text.replace("\\", "\\\\")
    text = text.replace("'", "''")

    return "'" + text + "'"


def value_sql(value, column):
    column = clean_name(column)

    if is_blank(value):
        return "NULL"

    if isinstance(value, str) and value.strip().upper() == "NULL":
        return "NULL"

    if column in DATE_COLUMNS:
        dt = to_datetime_value(value)
        if dt is None:
            return "NULL"
        return sql_string(dt.strftime("%Y-%m-%d"))

    if column in DATETIME_COLUMNS:
        dt = to_datetime_value(value)
        if dt is None:
            return "NULL"
        return sql_string(dt.strftime("%Y-%m-%d %H:%M:%S"))

    if column in TIME_COLUMNS:
        tm = to_time_value(value)
        if tm is None:
            return "NULL"
        return sql_string(tm.strftime("%H:%M:%S"))

    if isinstance(value, datetime.datetime):
        return sql_string(value.strftime("%Y-%m-%d %H:%M:%S"))

    if isinstance(value, datetime.date):
        return sql_string(value.strftime("%Y-%m-%d"))

    if isinstance(value, datetime.time):
        return sql_string(value.strftime("%H:%M:%S"))

    if isinstance(value, bool):
        return "1" if value else "0"

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    return sql_string(str(value).strip())


def read_headers(sheet):
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)

    if first_row is None:
        return []

    headers = list(first_row)

    while headers and headers[-1] is None:
        headers.pop()

    return [clean_name(h) for h in headers]


def check_headers(table, headers):
    expected = TABLE_COLUMNS[table]

    if headers != expected:
        raise ValueError(
            "Λάθος στήλες στο sheet "
            + table
            + "\nExcel: "
            + str(headers)
            + "\nInstall: "
            + str(expected)
        )


def read_rows(workbook, table):
    sheet = workbook[table]
    headers = read_headers(sheet)

    check_headers(table, headers)

    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row = list(row[:len(headers)])

        while len(row) < len(headers):
            row.append(None)

        if not all(is_blank(v) for v in row):
            rows.append(row)

    return headers, rows


def insert_multi(table, headers, rows):
    if not rows:
        return ""

    columns = ", ".join(f"`{h}`" for h in headers)

    values_lines = []

    for row in rows:
        values = ", ".join(
            value_sql(row[i] if i < len(row) else None, headers[i])
            for i in range(len(headers))
        )
        values_lines.append("(" + values + ")")

    return f"INSERT INTO `{table}` ({columns}) VALUES\n" + ",\n".join(values_lines) + ";\n"


def write_start(file):
    file.write("SET NAMES utf8mb4;\n")
    file.write(f"USE `{DATABASE_NAME}`;\n\n")
    file.write("START TRANSACTION;\n\n")


def write_end(file):
    file.write("COMMIT;\n")


def write_batches(file, table, headers, rows):
    total = 0

    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]
        file.write(insert_multi(table, headers, batch))
        file.write("\n")
        total += len(batch)

    return total


def write_doctor(file, workbook):
    headers, rows = read_rows(workbook, "doctor")

    supervisor_index = headers.index("supervisor_amka")
    amka_index = headers.index("amka")

    insert_headers = [h for h in headers if h != "supervisor_amka"]
    insert_indexes = [i for i, h in enumerate(headers) if h != "supervisor_amka"]

    insert_rows = []
    updates = []

    for row in rows:
        insert_rows.append([row[i] if i < len(row) else None for i in insert_indexes])

        supervisor = row[supervisor_index]
        amka = row[amka_index]

        if not is_blank(supervisor) and not is_blank(amka):
            updates.append((amka, supervisor))

    total = write_batches(file, "doctor", insert_headers, insert_rows)

    for amka, supervisor in updates:
        file.write(
            "UPDATE `doctor` SET `supervisor_amka` = "
            + value_sql(supervisor, "supervisor_amka")
            + " WHERE `amka` = "
            + value_sql(amka, "amka")
            + ";\n"
        )

    file.write("\n")

    return total


def write_table(file, workbook, table):
    if table == "doctor":
        return write_doctor(file, workbook)

    headers, rows = read_rows(workbook, table)

    total = write_batches(file, table, headers, rows)

    file.write("\n")

    return total


def write_file(workbook, filename, tables, output_folder):
    path = os.path.join(output_folder, filename)

    with open(path, "w", encoding="utf-8") as file:
        write_start(file)

        for table in tables:
            if table in workbook.sheetnames:
                count = write_table(file, workbook, table)
                print(filename, table, count)

        write_end(file)

    print("Created:", path)


# Γράφει συγκεκριμένες γραμμές ενός πίνακα.
# Το χρειαζόμαστε για να σπάσουμε το μεγάλο shift_assignment σε load10c, load10d, load10e.
def write_table_from_rows(file, table, headers, rows):
    total = write_batches(file, table, headers, rows)
    file.write("\n")
    return total


def write_custom_file(filename, table_data, output_folder):
    path = os.path.join(output_folder, filename)

    with open(path, "w", encoding="utf-8") as file:
        write_start(file)

        for table, headers, rows in table_data:
            count = write_table_from_rows(file, table, headers, rows)
            print(filename, table, count)

        write_end(file)

    print("Created:", path)


def split_in_three(rows):
    # Κόβει τις γραμμές σε 3 συνεχόμενα κομμάτια.
    # Δεν ανακατεύει τη σειρά, ώστε οι βάρδιες να φορτώνονται όπως είναι στο Excel.
    n = len(rows)
    cut1 = n // 3
    cut2 = (2 * n) // 3

    part1 = rows[:cut1]
    part2 = rows[cut1:cut2]
    part3 = rows[cut2:]

    return part1, part2, part3


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    excel_path = EXCEL_FILE

    if not os.path.isabs(excel_path):
        excel_path = os.path.join(base_dir, excel_path)

    if not os.path.exists(excel_path):
        alternative_path = EXCEL_FILE_ALTERNATIVE
        if not os.path.isabs(alternative_path):
            alternative_path = os.path.join(base_dir, alternative_path)

        if os.path.exists(alternative_path):
            excel_path = alternative_path
        else:
            raise FileNotFoundError("Δεν βρέθηκε το Excel: " + excel_path)

    output_folder = OUTPUT_FOLDER

    if not os.path.isabs(output_folder):
        output_folder = os.path.join(base_dir, output_folder)

    os.makedirs(output_folder, exist_ok=True)

    workbook = load_workbook(excel_path, read_only=True, data_only=True)

    write_file(workbook, "load10a.sql", [
        "staff",
        "doctor",
        "department",
        "nurse",
        "admin_staff",
        "doctor_department",
        "bed",
        "insurance",
        "patient",
        "emergency_contact",
        "diagnosis",
        "ken",
        "triage",
        "hospitalization",
        "exam",
        "operating_room",
        "medical_procedure",
        "procedure_staff",
    ], output_folder)

    write_file(workbook, "load10b.sql", [
        "drug",
        "active_substance",
        "drug_active_substance",
        "patient_allergy",
        "prescription",
    ], output_folder)

    # ------------------------------------------------------------------
    # Τα load10c, load10d, load10e είναι πλέον πιο ισορροπημένα.
    # Το μεγάλο sheet shift_assignment σπάει σε 3 κομμάτια.
    # Το load10c έχει πρώτα ΟΛΑ τα shift, ώστε τα shift_assignment
    # στα load10d και load10e να βρίσκουν ήδη τα αντίστοιχα shift_id.
    # ------------------------------------------------------------------

    shift_headers, shift_rows = read_rows(workbook, "shift")
    assignment_headers, assignment_rows = read_rows(workbook, "shift_assignment")

    assignment_part_1, assignment_part_2, assignment_part_3 = split_in_three(assignment_rows)

    evaluation_headers, evaluation_rows = read_rows(workbook, "hospitalization_evaluation")
    doctor_eval_headers, doctor_eval_rows = read_rows(workbook, "doctor_evaluation")
    image_headers, image_rows = read_rows(workbook, "entity_image")

    write_custom_file("load10c.sql", [
        ("shift", shift_headers, shift_rows),
        ("shift_assignment", assignment_headers, assignment_part_1),
    ], output_folder)

    write_custom_file("load10d.sql", [
        ("shift_assignment", assignment_headers, assignment_part_2),
        ("hospitalization_evaluation", evaluation_headers, evaluation_rows),
    ], output_folder)

    write_custom_file("load10e.sql", [
        ("shift_assignment", assignment_headers, assignment_part_3),
        ("doctor_evaluation", doctor_eval_headers, doctor_eval_rows),
        ("entity_image", image_headers, image_rows),
    ], output_folder)


if __name__ == "__main__":
    main()